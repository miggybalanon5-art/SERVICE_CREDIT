"""
Shared UI building blocks for the Form 6 Tracker: CSS injection, flash/toast
messaging, and small formatting/export helpers used by multiple tabs.

Nothing in this file touches the database directly - it's presentation-only,
so it can be imported by any tab module without pulling in write logic.
"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="0064E0")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LIGHT_BORDER = Border(
    left=Side(style="thin", color="CED0D4"),
    right=Side(style="thin", color="CED0D4"),
    top=Side(style="thin", color="CED0D4"),
    bottom=Side(style="thin", color="CED0D4"),
)

LEAVE_INTERNAL_COLS = ["id", "employee_id", "source_kind", "source_ref", "no_days", "no_halfdays"]
LEAVE_DISPLAY_RENAME = {
    "date_of_filing": "Date Filed", "month": "Month", "leave_type": "Leave Type",
    "total_days": "Days", "inclusive_dates": "Inclusive Dates",
    "service_credit_availed": "Credit Availed", "employee_label": "Employee", "grade": "Grade",
}
CREDIT_INTERNAL_COLS = ["id", "employee_id", "source_kind", "source_ref"]
CREDIT_DISPLAY_RENAME = {
    "credit_scope": "Scope", "event_date": "Date", "service_attended": "Service Attended",
    "credit_units": "Units", "employee_label": "Employee", "grade": "Grade",
}


def inject_app_css():
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

        html, body, [class*="css"] {
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif !important;
        }
        
        .block-container { 
            padding-top: 2rem; 
            padding-bottom: 2rem; 
            max-width: 1200px;
        }
        
        h1, h2, h3, h4, h5, h6 { font-weight: 600; }
        
        .stButton > button {
            background-color: var(--primary-color, #0064E0);
            color: #FFFFFF;
            border: none;
            border-radius: 6px;
            padding: 0.5rem 1rem;
            font-weight: 600;
            box-shadow: none;
            transition: opacity 0.2s;
            min-height: 44px;
        }
        .stButton > button:hover { opacity: 0.9; color: #FFFFFF; }
        .stButton > button:active { opacity: 0.8; color: #FFFFFF; }
        
        .stDownloadButton > button {
            background-color: var(--secondary-background-color);
            color: var(--text-color);
            border: none;
            border-radius: 6px;
            font-weight: 600;
            min-height: 44px;
        }
        .stDownloadButton > button:hover { opacity: 0.8; color: var(--text-color); }
        
        div[data-testid="stMetric"] {
            background-color: var(--background-color);
            border: 1px solid var(--secondary-background-color);
            border-radius: 8px;
            padding: 1.2rem;
            box-shadow: 0 1px 2px rgba(0, 0, 0, 0.05);
        }
        div[data-testid="stMetricValue"] {
            color: var(--primary-color, #0064E0);
            font-size: 2.2rem;
            font-weight: 700;
        }
        div[data-testid="stMetricLabel"] { font-size: 0.95rem; font-weight: 600; }
        
        /* 1. METRICS: FORCE HORIZONTAL SINGLE LINE FOR BOTH DESKTOP & MOBILE */
        div[data-testid="stHorizontalBlock"]:has(div[data-testid="stMetric"]) {
            display: flex !important;
            flex-direction: row !important;
            flex-wrap: nowrap !important;
            gap: 2px !important;
        }
        div[data-testid="stHorizontalBlock"]:has(div[data-testid="stMetric"]) > div[data-testid="column"] {
            width: auto !important;
            flex: 1 1 0px !important; 
        }

        /* 2. CALENDAR: FACEBOOK COVER PHOTO ASPECT RATIO (320x75) */
        iframe[title*="calendar" i],
        .fb-cover-calendar iframe,
        div[data-testid="stHtml"]:has(.fb-cover-calendar) {
            width: 100% !important;
            aspect-ratio: 820 / 312 !important;
            height: auto !important;
            object-fit: cover;
            border: 1px solid var(--secondary-background-color);
            border-radius: 8px;
        }
        
        .streamlit-expanderHeader { font-weight: 600; background-color: transparent; border-radius: 8px; }
        [data-testid="stSidebar"] { border-right: 1px solid var(--secondary-background-color); }
        
        .stTabs [data-baseweb="tab-list"] {
            gap: 20px;
            border-bottom: 1px solid var(--secondary-background-color);
            overflow-x: auto;
            overflow-y: hidden;
            flex-wrap: nowrap !important;
            -webkit-overflow-scrolling: touch;
            scrollbar-width: thin;
        }
        .stTabs [data-baseweb="tab"] {
            height: 50px;
            white-space: nowrap;
            background-color: transparent;
            border-radius: 0;
            font-weight: 600;
            flex-shrink: 0;
        }
        .stTabs [aria-selected="true"] { color: var(--primary-color, #0064E0) !important; border-bottom: 3px solid var(--primary-color, #0064E0) !important; }
        
        div[role="radiogroup"] { flex-direction: row; flex-wrap: wrap; gap: 12px; padding-bottom: 15px; margin-bottom: 20px; }
        div[role="radiogroup"] label { background-color: var(--background-color); border: 1px solid var(--secondary-background-color); border-radius: 20px; padding: 8px 18px; font-weight: 600; }
        div[role="radiogroup"] label[data-checked="true"] { background-color: var(--secondary-background-color); border-color: var(--primary-color, #0064E0); color: var(--primary-color, #0064E0); }
        
        [data-testid="stForm"] { background-color: var(--background-color); padding: 1.5rem; border-radius: 8px; border: 1px solid var(--secondary-background-color); box-shadow: 0 1px 2px rgba(0, 0, 0, 0.05); }
        [data-testid="stDataFrame"] { background-color: var(--background-color); border: 1px solid var(--secondary-background-color); border-radius: 8px; padding: 10px; box-shadow: 0 1px 2px rgba(0, 0, 0, 0.05); overflow-x: auto; }
        
        div[data-testid="stHorizontalBlock"] { padding: 6px 10px; border-radius: 6px; transition: background-color 0.15s ease-in-out; }
        div[data-testid="stHorizontalBlock"]:hover { background-color: var(--secondary-background-color) !important; }
        
        .small-note { font-size: 0.95rem; margin-bottom: 1rem; opacity: 0.8; }
        .section-divider { margin: 1.5rem 0; border-top: 1px solid var(--secondary-background-color); }

        div[data-baseweb="input"] > div,
        div[data-baseweb="textarea"] > div,
        div[data-baseweb="select"] > div,
        .stTextInput input,
        .stTextArea textarea,
        .stNumberInput input,
        .stDateInput input {
            border: 1px solid color-mix(in srgb, var(--text-color) 16%, transparent) !important;
            border-radius: 6px !important;
            transition: border-color 0.15s ease-in-out;
        }
        div[data-baseweb="input"] > div:focus-within,
        div[data-baseweb="textarea"] > div:focus-within,
        div[data-baseweb="select"] > div:focus-within {
            border-color: var(--primary-color, #0064E0) !important;
        }

        @media (max-width: 768px) {
            .block-container {
                padding-top: 1rem;
                padding-bottom: 1rem;
                max-width: 100%;
            }

            h1 { font-size: 1.6rem; }
            h2 { font-size: 1.35rem; }
            h3 { font-size: 1.15rem; }

            .stButton > button,
            .stDownloadButton > button {
                min-height: 48px;
                font-size: 1rem;
                padding: 0.75rem 1.25rem;
            }

            /* Stack standard blocks contextually, ignoring metrics via priority above */
            div[data-testid="stHorizontalBlock"] {
                flex-direction: column !important;
                width: 100% !important;
            }
            div[data-testid="stHorizontalBlock"] > div[data-testid="column"] {
                width: 100% !important;
            }

            div[data-testid="stMetric"] { padding: 0.85rem; }
            div[data-testid="stMetricValue"] { font-size: 1.6rem; }
            div[data-testid="stMetricLabel"] { font-size: 0.85rem; }

            .stTabs [data-baseweb="tab-list"] { gap: 10px; }
            .stTabs [data-baseweb="tab"] {
                height: 42px;
                font-size: 0.85rem;
                padding: 0 4px;
            }

            .streamlit-expanderHeader {
                padding: 1rem;
                font-size: 1.05rem;
            }

            [data-testid="stDataFrame"] { padding: 4px; }
        }

        @media (max-width: 480px) {
            h1 { font-size: 1.4rem; }
            h2 { font-size: 1.2rem; }

            .stButton > button,
            .stDownloadButton > button {
                min-height: 44px;
                font-size: 0.95rem;
                padding: 0.65rem 1rem;
            }

            div[data-testid="stMetricValue"] { font-size: 1.4rem; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def confirm_destructive_action(action_id: str, action_label: str, action_type: str = "delete") -> bool:
    confirm_key = f"_confirm_{action_id}"
    
    if st.session_state.get(confirm_key):
        col1, col2 = st.columns(2)
        with col1:
            st.warning(f"⚠️ This {action_type} cannot be undone!")
        with col2:
            if st.button("❌ Cancel", key=f"cancel_{action_id}", use_container_width=True):
                st.session_state[confirm_key] = False
                st.rerun()
        
        confirm_col, warning_col = st.columns([1, 1.5])
        with confirm_col:
            if st.button(f"✓ Confirm {action_type.title()}", key=f"confirm_{action_id}", 
                        use_container_width=True, type="primary"):
                st.session_state[confirm_key] = False
                return True
        
        with warning_col:
            st.caption(f"Click 'Confirm {action_type.title()}' to proceed.")
        
        return False
    else:
        st.session_state[confirm_key] = True
        st.warning(f"⚠️ Are you sure you want to {action_type} {action_label}? Click the button again to confirm.")
        st.rerun()


def keyed_tabs(labels: list[str], session_key: str):
    try:
        return st.tabs(labels, key=session_key)
    except TypeError:
        return st.tabs(labels)


def flash(message: str, kind: str = "success") -> None:
    st.session_state["flash_message"] = {"message": message, "kind": kind}


def show_flash() -> None:
    payload = st.session_state.pop("flash_message", None)
    if not payload: return
    kind = payload.get("kind", "success")
    message = payload.get("message", "")
    if kind == "error": st.error(message)
    elif kind == "warning": st.warning(message)
    else: st.success(message)


def queue_toast(message: str, icon: str = "✅") -> None:
    st.session_state["pending_toast"] = {"message": message, "icon": icon}


def show_pending_toast() -> None:
    payload = st.session_state.pop("pending_toast", None)
    if not payload: return
    st.toast(payload.get("message", ""), icon=payload.get("icon", "✅"))


def safe_text(value) -> str:
    if value is None: return ""
    if isinstance(value, float) and pd.isna(value): return ""
    if isinstance(value, (pd.Timestamp, datetime, date)):
        if pd.isna(value): return ""
        return pd.Timestamp(value).strftime("%Y-%m-%d")
    return str(value)


def readable_view(df: pd.DataFrame, drop: list[str] | None = None, rename: dict[str, str] | None = None) -> pd.DataFrame:
    if df.empty: return df.copy()
    drop = drop or []
    result = df.drop(columns=drop, errors="ignore").copy()
    for column in result.columns:
        if pd.api.types.is_datetime64_any_dtype(result[column]):
            result[column] = result[column].dt.strftime("%Y-%m-%d").fillna("")
    for column in result.columns:
        if pd.api.types.is_datetime64tz_dtype(result[column]):
            result[column] = result[column].dt.tz_convert(None).dt.strftime("%Y-%m-%d")
    if rename:
        result = result.rename(columns={k: v for k, v in rename.items() if k in result.columns})
    return result


def month_index_from_date(value: date | datetime | None) -> int:
    if value is None: return 0
    return max(0, min(11, pd.Timestamp(value).month - 1))


def date_bounds(leaves: pd.DataFrame, credits: pd.DataFrame) -> tuple[pd.Timestamp | None, pd.Timestamp | None]:
    if leaves.empty or "date_of_filing" not in leaves.columns: return None, None
    series = pd.to_datetime(leaves["date_of_filing"], errors="coerce").dropna()
    if series.empty: return None, None
    return series.min(), series.max()


def employee_options(df: pd.DataFrame) -> dict[str, int]:
    if df.empty: return {}
    return dict(zip(df["employee_label"].tolist(), df["id"].tolist(), strict=False))


def render_metrics(employees: pd.DataFrame, leaves: pd.DataFrame) -> None:
    registered_employees = len(employees)
    today_ts = pd.Timestamp(date.today())
    upcoming_leaves_count = 0
    leaves_only = leaves[leaves.get("source_kind", "") != "biometrics"] if not leaves.empty else leaves
    if not leaves_only.empty and "date_of_filing" in leaves_only.columns:
        parsed_dates = pd.to_datetime(leaves_only["date_of_filing"], errors="coerce")
        upcoming_leaves_count = int((parsed_dates >= today_ts).sum())

    cols = st.columns(2)
    cols[0].metric("Total Employees", f"{registered_employees:,}")
    cols[1].metric("Ongoing / Upcoming Leaves", f"{upcoming_leaves_count:,}")


def export_workbook_bytes(tables: dict[str, pd.DataFrame]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, df in tables.items():
        ws = workbook.create_sheet(title=title[:31])
        view = readable_view(df)
        if view.empty:
            ws.append(["No rows available."])
            continue
        ws.append([str(col) for col in view.columns])
        for row in view.itertuples(index=False, name=None): ws.append(list(row))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = LIGHT_BORDER
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = LIGHT_BORDER
                cell.alignment = Alignment(vertical="top")
        for col_idx, column_name in enumerate(view.columns, start=1):
            values = [str(v)[:80] for v in view.iloc[:, col_idx - 1].fillna("").tolist()]
            max_len = max([len(str(column_name))] + [len(v) for v in values]) if values else len(str(column_name))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 34)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def db_backup_bytes() -> bytes:
    from form6_store import DB_PATH, ensure_database
    ensure_database()
    return DB_PATH.read_bytes()
