from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import re
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


MONTHS = [
    "JANUARY",
    "FEBRUARY",
    "MARCH",
    "APRIL",
    "MAY",
    "JUNE",
    "JULY",
    "AUGUST",
    "SEPTEMBER",
    "OCTOBER",
    "NOVEMBER",
    "DECEMBER",
]

LEAVE_COLUMNS = [
    "record_id",
    "grade",
    "employee_name",
    "employee_sheet",
    "last_name",
    "first_name",
    "middle_initial",
    "position",
    "date_of_filing",
    "date_of_filing_raw",
    "inclusive_dates",
    "month",
    "month_raw",
    "no_days",
    "no_halfdays",
    "total_days",
    "leave_type",
    "service_credit_availed",
    "local_balance",
    "national_balance",
    "source_file",
    "source_sheet",
    "source_row",
]

SUMMARY_COLUMNS = [
    "grade",
    "employee_name",
    "employee_sheet",
    "position",
    "leave_entries",
    "leave_days",
    "personal_leave",
    "sick_leave",
    "vacation_leave",
    "special_privilege_leave",
    "other_leave",
    "local_earned",
    "local_used",
    "local_balance",
    "national_earned",
    "national_used",
    "national_balance",
    "last_filing_date",
    "source_file",
]

CREDIT_COLUMNS = [
    "grade",
    "employee_name",
    "employee_sheet",
    "credit_scope",
    "event_date",
    "event_date_raw",
    "service_attended",
    "credit_units",
    "source_file",
    "source_sheet",
    "source_row",
]


@dataclass(frozen=True)
class ConsolidatedForm6:
    leave_entries: pd.DataFrame
    employee_summary: pd.DataFrame
    service_credits: pd.DataFrame
    import_log: pd.DataFrame


def default_workbook_paths(data_dir: Path | None = None) -> list[Path]:
    base = data_dir or Path(__file__).with_name("data")
    return sorted(base.glob("*FORM 6*.xlsx"), key=lambda path: grade_sort_key(path.name))


def consolidate_workbooks(sources: Iterable[tuple[str, bytes | str | Path]]) -> ConsolidatedForm6:
    leave_records: list[dict] = []
    employee_records: list[dict] = []
    credit_records: list[dict] = []
    log_records: list[dict] = []

    for source_name, payload in sources:
        try:
            workbook_data = read_form6_workbook(source_name, payload)
        except Exception as exc:  # noqa: BLE001 - import log should capture workbook issues.
            log_records.append(
                {
                    "source_file": source_name,
                    "status": "error",
                    "employee_sheets": 0,
                    "leave_entries": 0,
                    "service_credit_rows": 0,
                    "message": str(exc),
                }
            )
            continue

        leave_records.extend(workbook_data["leave_records"])
        employee_records.extend(workbook_data["employee_records"])
        credit_records.extend(workbook_data["credit_records"])
        log_records.append(
            {
                "source_file": source_name,
                "status": "loaded",
                "employee_sheets": workbook_data["employee_sheets"],
                "leave_entries": len(workbook_data["leave_records"]),
                "service_credit_rows": len(workbook_data["credit_records"]),
                "message": "",
            }
        )

    leave_df = pd.DataFrame(leave_records, columns=LEAVE_COLUMNS)
    summary_df = pd.DataFrame(employee_records, columns=SUMMARY_COLUMNS)
    credits_df = pd.DataFrame(credit_records, columns=CREDIT_COLUMNS)
    log_df = pd.DataFrame(log_records)

    leave_df = finalise_leave_entries(leave_df)
    credits_df = finalise_service_credits(credits_df)
    summary_df = finalise_employee_summary(summary_df, leave_df)

    return ConsolidatedForm6(
        leave_entries=leave_df,
        employee_summary=summary_df,
        service_credits=credits_df,
        import_log=log_df,
    )


def read_form6_workbook(source_name: str, payload: bytes | str | Path) -> dict:
    workbook = load_workbook(open_workbook_payload(payload), data_only=True, read_only=False)
    grade = grade_from_filename(source_name)

    leave_records: list[dict] = []
    employee_records: list[dict] = []
    credit_records: list[dict] = []
    employee_sheet_count = 0

    for sheet in workbook.worksheets:
        header_row = find_tracker_header_row(sheet)
        if not header_row or not is_employee_sheet(sheet.title):
            continue

        employee_sheet_count += 1
        employee_context = read_employee_context(sheet, grade, source_name, header_row)
        employee_records.append(employee_context)
        leave_records.extend(read_leave_entries(sheet, employee_context, header_row))
        credit_records.extend(read_service_credit_entries(sheet, employee_context))

    workbook.close()
    return {
        "employee_sheets": employee_sheet_count,
        "leave_records": leave_records,
        "employee_records": employee_records,
        "credit_records": credit_records,
    }


def open_workbook_payload(payload: bytes | str | Path):
    if isinstance(payload, bytes):
        return BytesIO(payload)
    return Path(payload)


def find_tracker_header_row(sheet) -> int | None:
    for row_idx in range(1, min(sheet.max_row, 12) + 1):
        row_values = [clean_text(sheet.cell(row_idx, col_idx).value).upper() for col_idx in range(1, 15)]
        has_name = "LAST NAME" in row_values and any("FIRST" in value for value in row_values)
        has_date = any("DATE OF FILING" in value for value in row_values)
        has_leave = any("TYPE OF LEAVE" in value for value in row_values)
        if has_name and has_date and has_leave:
            return row_idx
    return None


def is_employee_sheet(sheet_name: str) -> bool:
    normalized = sheet_name.strip().upper()
    return normalized not in {"2025", "SHEET1"}


def read_employee_context(sheet, grade: str, source_name: str, header_row: int) -> dict:
    full_name = clean_text(sheet["A1"].value)
    last_name = first_non_empty(sheet, 2, header_row + 1, header_row + 25)
    first_name = first_non_empty(sheet, 3, header_row + 1, header_row + 25)
    middle_initial = first_non_empty(sheet, 4, header_row + 1, header_row + 25)
    position = first_non_empty(sheet, 5, header_row + 1, min(sheet.max_row, header_row + 120))
    employee_name = full_name or compose_employee_name(last_name, first_name, middle_initial) or sheet.title

    return {
        "grade": grade,
        "employee_name": employee_name,
        "employee_sheet": sheet.title,
        "position": position,
        "leave_entries": 0,
        "leave_days": 0.0,
        "personal_leave": as_number(sheet.cell(16, 23).value),
        "sick_leave": as_number(sheet.cell(16, 25).value),
        "vacation_leave": as_number(sheet.cell(16, 27).value),
        "special_privilege_leave": as_number(sheet.cell(16, 29).value),
        "other_leave": as_number(sheet.cell(16, 31).value),
        "local_earned": as_number(sheet.cell(23, 30).value),
        "local_used": as_number(sheet.cell(16, 33).value),
        "local_balance": as_number(sheet.cell(17, 33).value),
        "national_earned": as_number(sheet.cell(23, 21).value),
        "national_used": as_number(sheet.cell(16, 35).value),
        "national_balance": as_number(sheet.cell(17, 35).value),
        "last_filing_date": pd.NaT,
        "source_file": source_name,
    }


def read_leave_entries(sheet, employee_context: dict, header_row: int) -> list[dict]:
    records: list[dict] = []
    current_last = clean_text(sheet.cell(header_row + 1, 2).value)
    current_first = clean_text(sheet.cell(header_row + 1, 3).value)
    current_middle = clean_text(sheet.cell(header_row + 1, 4).value)
    current_position = clean_text(sheet.cell(header_row + 1, 5).value) or employee_context["position"]

    for row_idx in range(header_row + 1, min(sheet.max_row, 1000) + 1):
        row_no = clean_text(sheet.cell(row_idx, 1).value)
        last_name = clean_text(sheet.cell(row_idx, 2).value)
        first_name = clean_text(sheet.cell(row_idx, 3).value)
        middle_initial = clean_text(sheet.cell(row_idx, 4).value)
        position = clean_text(sheet.cell(row_idx, 5).value)
        date_of_filing_raw = clean_text(sheet.cell(row_idx, 6).value)
        inclusive_dates = clean_text(sheet.cell(row_idx, 7).value)
        month_raw = clean_text(sheet.cell(row_idx, 8).value)
        no_days = as_number(sheet.cell(row_idx, 9).value)
        no_halfdays_raw = as_number(sheet.cell(row_idx, 10).value)
        no_halfdays = half_days_to_day_equivalent(no_halfdays_raw)
        leave_type = clean_text(sheet.cell(row_idx, 11).value)
        service_credit_availed = clean_text(sheet.cell(row_idx, 12).value)
        local_balance = as_number(sheet.cell(row_idx, 13).value)
        national_balance = as_number(sheet.cell(row_idx, 14).value)

        if last_name:
            current_last = last_name
        if first_name:
            current_first = first_name
        if middle_initial:
            current_middle = middle_initial
        if position:
            current_position = position

        has_entry_signal = any(
            [
                date_of_filing_raw,
                inclusive_dates,
                month_raw,
                no_days,
                no_halfdays,
                leave_type,
                service_credit_availed,
            ]
        )
        if not has_entry_signal:
            continue

        records.append(
            {
                "record_id": f"{employee_context['grade']}|{employee_context['employee_sheet']}|{row_idx}",
                "grade": employee_context["grade"],
                "employee_name": employee_context["employee_name"],
                "employee_sheet": employee_context["employee_sheet"],
                "last_name": current_last,
                "first_name": current_first,
                "middle_initial": current_middle,
                "position": current_position,
                "date_of_filing": parse_single_date(date_of_filing_raw),
                "date_of_filing_raw": date_of_filing_raw,
                "inclusive_dates": inclusive_dates,
                "month": normalize_month(month_raw),
                "month_raw": month_raw,
                "no_days": no_days,
                "no_halfdays": no_halfdays,
                "total_days": compute_total_days(no_days, no_halfdays),
                "leave_type": normalize_leave_type(leave_type),
                "service_credit_availed": normalize_service_credit(service_credit_availed),
                "local_balance": local_balance,
                "national_balance": national_balance,
                "source_file": employee_context["source_file"],
                "source_sheet": sheet.title,
                "source_row": row_idx,
            }
        )

    return records


def read_service_credit_entries(sheet, employee_context: dict) -> list[dict]:
    records: list[dict] = []
    for row_idx in range(25, min(sheet.max_row, 1000) + 1):
        national_event = clean_text(sheet.cell(row_idx, 19).value)
        national_units = as_number(sheet.cell(row_idx, 21).value)
        if national_units > 0:
            records.append(
                {
                    "grade": employee_context["grade"],
                    "employee_name": employee_context["employee_name"],
                    "employee_sheet": employee_context["employee_sheet"],
                    "credit_scope": "DIVISION",
                    "event_date": pd.NaT,
                    "event_date_raw": "",
                    "service_attended": national_event,
                    "credit_units": national_units,
                    "source_file": employee_context["source_file"],
                    "source_sheet": sheet.title,
                    "source_row": row_idx,
                }
            )

        local_date_raw = clean_text(sheet.cell(row_idx, 27).value)
        local_event = clean_text(sheet.cell(row_idx, 28).value)
        local_units = as_number(sheet.cell(row_idx, 30).value)
        if local_units > 0:
            records.append(
                {
                    "grade": employee_context["grade"],
                    "employee_name": employee_context["employee_name"],
                    "employee_sheet": employee_context["employee_sheet"],
                    "credit_scope": "LOCAL",
                    "event_date": parse_single_date(local_date_raw),
                    "event_date_raw": local_date_raw,
                    "service_attended": local_event,
                    "credit_units": local_units,
                    "source_file": employee_context["source_file"],
                    "source_sheet": sheet.title,
                    "source_row": row_idx,
                }
            )
    return records


def finalise_leave_entries(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=LEAVE_COLUMNS)
    df = df.copy()
    df["date_of_filing"] = pd.to_datetime(df["date_of_filing"], errors="coerce")
    numeric_columns = ["no_days", "no_halfdays", "total_days", "local_balance", "national_balance"]
    for column in numeric_columns:
        df[column] = pd.to_numeric(df[column], errors="coerce").fillna(0)
    df["_grade_sort"] = df["grade"].map(grade_sort_key)
    df = df.sort_values(["_grade_sort", "employee_name", "date_of_filing", "source_row"], na_position="last")
    return df.drop(columns=["_grade_sort"]).reset_index(drop=True)


def finalise_service_credits(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=CREDIT_COLUMNS)
    df = df.copy()
    df["event_date"] = pd.to_datetime(df["event_date"], errors="coerce")
    df["credit_units"] = pd.to_numeric(df["credit_units"], errors="coerce").fillna(0)
    df["_grade_sort"] = df["grade"].map(grade_sort_key)
    df = df.sort_values(["_grade_sort", "employee_name", "credit_scope", "event_date", "source_row"], na_position="last")
    return df.drop(columns=["_grade_sort"]).reset_index(drop=True)


def finalise_employee_summary(summary_df: pd.DataFrame, leave_df: pd.DataFrame) -> pd.DataFrame:
    if summary_df.empty:
        return pd.DataFrame(columns=SUMMARY_COLUMNS)

    summary_df = summary_df.copy()
    numeric_columns = [
        "leave_entries",
        "leave_days",
        "personal_leave",
        "sick_leave",
        "vacation_leave",
        "special_privilege_leave",
        "other_leave",
        "local_earned",
        "local_used",
        "local_balance",
        "national_earned",
        "national_used",
        "national_balance",
    ]
    for column in numeric_columns:
        summary_df[column] = pd.to_numeric(summary_df[column], errors="coerce").fillna(0)

    if not leave_df.empty:
        grouped = (
            leave_df.groupby(["grade", "employee_sheet"], dropna=False)
            .agg(
                leave_entries=("record_id", "count"),
                leave_days=("total_days", "sum"),
                last_filing_date=("date_of_filing", "max"),
            )
            .reset_index()
        )
        summary_df = summary_df.drop(columns=["leave_entries", "leave_days", "last_filing_date"]).merge(
            grouped,
            on=["grade", "employee_sheet"],
            how="left",
        )
        summary_df["leave_entries"] = summary_df["leave_entries"].fillna(0).astype(int)
        summary_df["leave_days"] = summary_df["leave_days"].fillna(0)
    else:
        summary_df["last_filing_date"] = pd.NaT

    summary_df["last_filing_date"] = pd.to_datetime(summary_df["last_filing_date"], errors="coerce")
    summary_df["_grade_sort"] = summary_df["grade"].map(grade_sort_key)
    summary_df = summary_df.sort_values(["_grade_sort", "employee_name"]).drop(columns=["_grade_sort"])
    return summary_df.reset_index(drop=True)


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, (datetime, date)):
        return f"{value.strftime('%B')} {value.day}, {value.year}"
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def half_days_to_day_equivalent(raw: float) -> float:
    """Convert Form 6 half-day values to day-equivalent storage.

    Excel workbooks usually store a *count* of half-day periods (1, 2, …).
    Manual encoding in this app uses day-equivalent values (0.5, 1.0, …).
    """
    value = float(raw or 0)
    if value <= 0:
        return 0.0
    if value == int(value) and value >= 1:
        return value * 0.5
    return value


def compute_total_days(no_days: float, no_halfdays: float) -> float:
    """Total leave days = whole days plus half-day equivalent."""
    return float(no_days or 0) + float(no_halfdays or 0)


def month_from_date_value(value) -> str:
    """Derive the Form 6 month label from a filing date."""
    parsed = parse_single_date(value)
    if pd.isna(parsed):
        return ""
    return MONTHS[pd.Timestamp(parsed).month - 1]


def as_number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value).replace(",", "")
    if text in {"", "-"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_single_date(value) -> pd.Timestamp:
    if value is None or value == "":
        return pd.NaT
    if isinstance(value, (datetime, date)):
        return pd.to_datetime(value, errors="coerce")

    text = clean_text(value)
    if text in {"", "-"}:
        return pd.NaT
    text = re.sub(r"\s+\.", ".", text)
    text = re.sub(r"\bSept\.?\b", "Sep", text, flags=re.IGNORECASE)
    text = re.sub(r"\b([A-Za-z]{3,9})\.", r"\1", text)
    return pd.to_datetime(text, errors="coerce")


def normalize_month(value) -> str:
    text = clean_text(value).upper()
    if not text:
        return ""
    compact = re.sub(r"[^A-Z]", "", text)
    corrections = {
        "JANAURY": "JANUARY",
        "JAN": "JANUARY",
        "FEB": "FEBRUARY",
        "FEBRAURY": "FEBRUARY",
        "MAR": "MARCH",
        "APR": "APRIL",
        "AUG": "AUGUST",
        "SEPT": "SEPTEMBER",
        "SEP": "SEPTEMBER",
        "OCT": "OCTOBER",
        "NOV": "NOVEMBER",
        "DEC": "DECEMBER",
    }
    if compact in corrections:
        return corrections[compact]
    for month in MONTHS:
        if compact.startswith(month[:3]):
            return month
    return text


def normalize_leave_type(value) -> str:
    text = clean_text(value).upper()
    text = text.replace("LAEVE", "LEAVE").replace("PRIVILEDGE", "PRIVILEGE")
    text = re.sub(r"\s+", " ", text).strip()
    if text == "PERSONAL":
        return "PERSONAL LEAVE"
    if text == "SICK":
        return "SICK LEAVE"
    return text


def normalize_service_credit(value) -> str:
    text = clean_text(value).upper()
    if text in {"DIVISION", "NATIONAL"}:
        return "DIVISION"
    if text == "LOCAL":
        return "LOCAL"
    return text


def first_non_empty(sheet, column: int, start_row: int, end_row: int) -> str:
    for row_idx in range(start_row, min(sheet.max_row, end_row) + 1):
        value = clean_text(sheet.cell(row_idx, column).value)
        if value:
            return value
    return ""


def compose_employee_name(last_name: str, first_name: str, middle_initial: str) -> str:
    pieces = [piece for piece in [first_name, middle_initial] if piece]
    if last_name and pieces:
        return f"{last_name}, {' '.join(pieces)}"
    return last_name or " ".join(pieces)


def grade_from_filename(filename: str) -> str:
    stem = Path(filename).stem.upper().strip()
    if stem.startswith("NT"):
        return "NT"
    match = re.search(r"G\s*(\d{1,2})", stem)
    if match:
        return f"G{int(match.group(1))}"
    return stem.split("-")[0].strip() or "UNSPECIFIED"


def grade_sort_key(value: str | Path) -> int:
    text = str(value).upper()
    match = re.search(r"G\s*(\d{1,2})", text)
    if match:
        return int(match.group(1))
    if "NT" in text:
        return 99
    return 100


def month_sort_key(month: str) -> int:
    try:
        return MONTHS.index(str(month).upper())
    except ValueError:
        return len(MONTHS)
